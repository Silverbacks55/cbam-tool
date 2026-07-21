#!/usr/bin/env python3
"""
Rebuild the CBAM tool's data files from the Commission's official Excel files.

Usage:
    pip install openpyxl
    python3 build_data.py \
        --assessment "CBAM Self Assessment Tool Version 1.1.xlsx" \
        --benchmarks "CBAM Benchmarks_20260206.xlsx" \
        --dvs        "DVs as adopted_v20260204 .xlsx"

Outputs (overwrites):
    data/assessment-data.json   (CN codes, reporting fields, countries)
    data/cost-data.json         (benchmarks, default values)

When the Commission publishes updated files, download them from
https://taxation-customs.ec.europa.eu/carbon-border-adjustment-mechanism/cbam-legislation-and-guidance_en
and re-run this script. No code changes needed unless the file layouts change.
"""
import argparse, json, re, sys
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

def clean(v):
    if v is None: return ""
    return str(v).replace('\x02','').replace('_x0002_','').strip()

def build_assessment(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['CN Codes']
    headers = [clean(c.value) for c in ws[1]]
    keep = ['Main Category','Aggregated Category','CN Code','Goods concerned','CBAM Applies',
            'Indirect Emissions','Quantity','Country','Installation data','Special provisions',
            'Production Routes','Precursors','Extra','Indirect Emissions Data','Data quality',
            'Carbon Price Abroad']
    idx = {h: headers.index(h) for h in keep}
    pool, pidx = [], {}
    def sid(s):
        if s not in pidx:
            pidx[s] = len(pool); pool.append(s)
        return pidx[s]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx['CN Code']] is None: continue
        raw = str(row[idx['CN Code']]).strip()
        code = raw.zfill(8) if raw.isdigit() else raw
        rows.append([sid(code if h == 'CN Code' else clean(row[idx[h]])) for h in keep])
    ws = wb['Country Codes']
    countries = sorted(
        ([clean(r[1]), 1 if clean(r[2]) == "Yes" else 0]
         for r in ws.iter_rows(min_row=2, values_only=True) if r[1] is not None),
        key=lambda x: x[0])
    return {"fields": keep, "pool": pool, "rows": rows, "countries": countries}

def build_benchmarks(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['Benchmarks']
    bench, cur = {}, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        cn, desc, a, ra, b, rb = (list(row) + [None]*6)[:6]
        if isinstance(cn, str) and desc is None and a is None:  # section header
            continue
        if cn is not None:
            cur = re.sub(r'\D', '', str(int(cn)) if isinstance(cn, (int, float)) else str(cn)).zfill(8)
            bench[cur] = []
        if cur is None or (a is None and b is None): continue
        route = (ra or rb or '')
        route = route.strip() if isinstance(route, str) else ''
        bench[cur].append([route,
                           a if isinstance(a, (int, float)) else None,
                           b if isinstance(b, (int, float)) else None])
    wb.close()
    return bench

# Country-sheet names in the DV file -> country names in the assessment tool.
ALIASES = {
    'congo': ['Congo, Republic of'],
    'democraticrepublicofthecongo': ['Congo, Democratic Republic of'],
    'northkorea': ["Korea, Democratic People's Republic of"],
    'southkorea': ['Korea, Republic of'],
    'unitedkingdom': ['United Kingdom (Northern Ireland)',
                      'United Kingdom (excluding Northern Ireland)'],
    'unitedstates': ['United States of America'],
    'myanmarburma': ['Myanmar'],
    'cotedivoire': ["Côte d'Ivoire"],
}

def build_dvs(path, tool_countries):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rawcn = lambda s: re.sub(r'\D', '', str(s))
    num = lambda v: round(float(v), 4) if isinstance(v, (int, float)) else None
    def parse_sheet(ws):
        fullname, rows = None, []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row = (list(row) + [None]*9)[:9]
            if i == 0: fullname = row[0]; continue
            cn, desc = row[0], row[1]
            if cn in (None, 'Product CN Code') or desc is None: continue
            if not re.search(r'\d', str(cn)): continue
            rt = row[8] if isinstance(row[8], str) else ''
            rows.append((rawcn(cn), str(desc).strip(), num(row[2]), num(row[3]),
                         rt.replace('\xa0', '').strip()))
        return fullname, rows

    fb_sheet = next(s for s in wb.sheetnames if s.startswith('_Other'))
    _, fb_rows = parse_sheet(wb[fb_sheet])
    prod_index, products, fallback = {}, [], []
    for cn, desc, dd, ii, rt in fb_rows:
        prod_index[(cn, desc)] = len(products)
        products.append([cn, desc, rt]); fallback.append([dd, ii])

    nrm = lambda s: re.sub(r'[^a-z]', '', s.lower())
    by_norm = {nrm(n): n for n in tool_countries}
    sheets, vals, cmap, unmatched = [], [], {}, []
    for sheet in wb.sheetnames:
        if sheet in ('Overview', 'Version History') or sheet == fb_sheet: continue
        fullname, rows = parse_sheet(wb[sheet])
        fullname = (fullname or sheet).strip()
        key = nrm(fullname)
        targets = ALIASES.get(key) or ([by_norm[key]] if key in by_norm else None)
        if not targets:
            cands = [n for k, n in by_norm.items() if key in k or k in key]
            targets = [cands[0]] if len(cands) == 1 else None
        if not targets:
            unmatched.append(fullname); continue
        arr = [0]*len(products)
        for cn, desc, dd, ii, rt in rows:
            idx = prod_index.get((cn, desc))
            if idx is not None and (dd is not None or ii is not None):
                arr[idx] = [dd, ii]
        si = len(sheets); sheets.append(fullname); vals.append(arr)
        for t in targets: cmap[t] = si
    wb.close()
    if unmatched:
        print("WARNING - country sheets not matched (add to ALIASES):", unmatched)
    return products, fallback, sheets, vals, cmap

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--assessment', required=True)
    ap.add_argument('--benchmarks', required=True)
    ap.add_argument('--dvs', required=True)
    a = ap.parse_args()

    assessment = build_assessment(a.assessment)
    bench = build_benchmarks(a.benchmarks)
    products, fallback, sheets, vals, cmap = build_dvs(
        a.dvs, [n for n, _ in assessment['countries']])

    with open('data/assessment-data.json', 'w', encoding='utf8') as f:
        json.dump(assessment, f, separators=(',', ':'), ensure_ascii=False)
    with open('data/cost-data.json', 'w', encoding='utf8') as f:
        json.dump({'bench': bench, 'products': products, 'fallback': fallback,
                   'sheets': sheets, 'vals': vals, 'cmap': cmap},
                  f, separators=(',', ':'), ensure_ascii=False)

    codes = {assessment['pool'][r[assessment['fields'].index('CN Code')]]
             for r in assessment['rows']}
    keys = {p[0] for p in products}
    cov = sum(1 for c in codes if c in keys or c[:6] in keys or c[:4] in keys)
    print(f"CN codes: {len(codes)} | benchmark coverage: {len(codes & set(bench))} "
          f"| default-value coverage: {cov} | country tables: {len(sheets)}")
    print("Wrote data/assessment-data.json and data/cost-data.json")

if __name__ == '__main__':
    main()
